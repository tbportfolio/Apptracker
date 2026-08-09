"""Shared openpyxl helpers used across report, validation, and dashboard modules."""

from openpyxl.utils import get_column_letter


def autosize_columns(ws, start_col=1, end_col=None):

    if end_col is None:
        end_col = ws.max_column

    for col in range(start_col, end_col + 1):

        column_letter = get_column_letter(col)
        max_length = 0

        for cell in ws[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2
