"""Build Apptracker.xlsx and the Active/Rejected CSV exports."""

import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from xlsx_utils import autosize_columns
from dashboards import create_apptracker_dashboard


def build_report(active_applications, rejected_applications, config):
    """
    Create Apptracker.xlsx (Summary, Active, Rejected, Dashboard sheets)
    inside config.apptracker_folder. Returns the path to the saved file.
    """

    active = len(active_applications)
    rejected = len(rejected_applications)
    total = active + rejected
    active_percent = (active / total * 100) if total > 0 else 0
    rejected_percent = (rejected / total * 100) if total > 0 else 0

    config.apptracker_folder.mkdir(parents=True, exist_ok=True)

    excel_file = config.apptracker_folder / "Apptracker.xlsx"

    # Delete previous workbook if it exists
    if excel_file.exists():
        excel_file.unlink()

    wb = Workbook()

    # ==========================
    # Summary Sheet
    # ==========================
    summary = wb.active
    summary.title = "Summary"

    summary["A1"] = "Job Application Summary"

    summary["A3"] = "Active Applications"
    summary["B3"] = active

    summary["A4"] = "Rejected Applications"
    summary["B4"] = rejected

    summary["A5"] = "Total Submitted"
    summary["B5"] = total

    summary["A7"] = "Active %"
    summary["B7"] = active_percent / 100

    summary["A8"] = "Rejected %"
    summary["B8"] = rejected_percent / 100

    summary["A10"] = "Report Generated"
    summary["B10"] = datetime.now()

    summary["A12"] = "Motivation"
    summary["B12"] = (
        "Keep going! Most of your applications are still active."
        if active > rejected
        else "Time to submit some more applications."
    )

    summary.merge_cells("A1:B1")

    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 42

    summary["A1"].font = Font(size=16, bold=True)
    summary["A1"].alignment = Alignment(horizontal="center")

    summary["B7"].number_format = "0.0%"
    summary["B8"].number_format = "0.0%"

    summary["B10"].number_format = "mm/dd/yyyy hh:mm AM/PM"

    summary["B12"].alignment = Alignment(horizontal="right")

    # ==========================
    # Active Sheet
    # ==========================
    ws = wb.create_sheet("Active")

    ws.append(["Work Type", "Company", "Position", "Date", "Created Timestamp"])

    for app in active_applications:

        ws.append([
            app["work_type"],
            app["company"],
            app["position"],
            app["date"],
            app["created_timestamp"]
        ])

    # ==========================
    # Rejected Sheet
    # ==========================
    ws = wb.create_sheet("Rejected")

    ws.append(["Work Type", "Company", "Position", "Date", "Created Timestamp"])

    for app in rejected_applications:

        ws.append([
            app["work_type"],
            app["company"],
            app["position"],
            app["date"],
            app["created_timestamp"]
        ])

    # ==========================================
    # Format Workbook
    # ==========================================
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for ws in wb.worksheets:

        # Skip Summary sheet
        if ws.title == "Summary":
            continue

        # Bold blue headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Format and hide Created Timestamp column
        if ws.title in ("Active", "Rejected"):

            for row in ws.iter_rows(min_row=2):
                row[4].number_format = "yyyy-mm-dd hh:mm:ss"

            ws.column_dimensions["E"].hidden = True

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-size columns
        autosize_columns(ws)

        # Create an Excel Table
        last_row = ws.max_row
        last_column = get_column_letter(ws.max_column)

        table = Table(
            displayName=f"{ws.title}Table",
            ref=f"A1:{last_column}{last_row}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style
        ws.add_table(table)

    # Create the Dashboard sheet
    create_apptracker_dashboard(wb)

    # Save workbook
    wb.save(excel_file)

    print(f"Excel report saved to:\n{excel_file}\n")

    return excel_file


def export_csv(active_applications, rejected_applications, config):
    """
    Write Active.csv and Rejected.csv into config.apptracker_folder.
    Returns (active_csv, rejected_csv) paths.
    """

    active_csv = config.apptracker_folder / "Active.csv"
    rejected_csv = config.apptracker_folder / "Rejected.csv"

    # Delete previous CSV files if they exist
    if active_csv.exists():
        active_csv.unlink()

    if rejected_csv.exists():
        rejected_csv.unlink()

    # ==========================
    # Active.csv
    # ==========================
    with open(active_csv, "w", newline="", encoding="utf-8") as file:

        writer = csv.writer(file)

        writer.writerow(["Work Type", "Company", "Position", "Date", "Created Timestamp"])

        for app in active_applications:
            writer.writerow([
                app["work_type"],
                app["company"],
                app["position"],
                app["date"],
                app["created_timestamp"]
            ])

    # ==========================
    # Rejected.csv
    # ==========================
    with open(rejected_csv, "w", newline="", encoding="utf-8") as file:

        writer = csv.writer(file)

        writer.writerow(["Work Type", "Company", "Position", "Date", "Created Timestamp"])

        for app in rejected_applications:
            writer.writerow([
                app["work_type"],
                app["company"],
                app["position"],
                app["date"],
                app["created_timestamp"]
            ])

    print(f"CSV files saved to:\n{config.apptracker_folder}\n")

    return active_csv, rejected_csv
