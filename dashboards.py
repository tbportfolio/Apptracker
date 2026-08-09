"""Dashboard worksheet builders for Apptracker.xlsx and Validation.xlsx."""

from collections import Counter
from datetime import datetime

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

from xlsx_utils import autosize_columns


def create_apptracker_dashboard(workbook):
    """
    Creates the Dashboard worksheet for Apptracker.xlsx.
    Uses ONLY the Active worksheet.
    """

    # Remove existing Dashboard if it exists
    if "Dashboard" in workbook.sheetnames:
        del workbook["Dashboard"]

    ws_active = workbook["Active"]
    ws_dash = workbook.create_sheet("Dashboard")

    #
    # Dashboard Title
    #

    active_count = ws_active.max_row - 1

    ws_dash.merge_cells("A1:B1")

    ws_dash["A1"] = "Apptracker Dashboard"
    ws_dash["A1"].font = Font(size=16, bold=True)
    ws_dash["A1"].alignment = Alignment(horizontal="center")

    ws_dash["A3"] = "Active Applications"
    ws_dash["B3"] = active_count
    ws_dash["B3"].font = Font(size=16, bold=True)

    #
    # Read Active Sheet
    #

    worktype_counter = Counter()
    month_counter = Counter()
    newest = []

    for row in ws_active.iter_rows(min_row=2, values_only=True):

        work_type, company, position, date, created_timestamp = row

        if work_type:
            worktype_counter[work_type] += 1

        if date:
            # Internal key keeps months sorted correctly forever
            date_obj = datetime.strptime(date, "%m/%d/%Y")
            month_key = date_obj.strftime("%Y-%m")
            month_counter[month_key] += 1

        newest.append((created_timestamp, work_type, company, position, date))

    newest.sort(reverse=True)

    #
    # Pie Chart Data
    #

    ws_dash["A6"] = "Work Type"
    ws_dash["B6"] = "Count"

    row = 7

    for work_type, count in worktype_counter.items():
        ws_dash.cell(row=row, column=1).value = work_type
        ws_dash.cell(row=row, column=2).value = count
        row += 1

    pie = PieChart()

    labels = Reference(
        ws_dash,
        min_col=1,
        min_row=7,
        max_row=row - 1
    )

    data = Reference(
        ws_dash,
        min_col=2,
        min_row=6,
        max_row=row - 1
    )

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Applications by Work Type"

    ws_dash.add_chart(pie, "D3")

    #
    # Monthly Totals
    #

    ws_dash["A15"] = "Month"
    ws_dash["B15"] = "Applications"

    months = sorted(month_counter.keys())

    row = 16

    for month in months:

        display_month = datetime.strptime(
            month,
            "%Y-%m"
        ).strftime("%b %Y")

        ws_dash.cell(row=row, column=1).value = display_month
        ws_dash.cell(row=row, column=2).value = month_counter[month]

        row += 1

    bar = BarChart()

    data = Reference(
        ws_dash,
        min_col=2,
        min_row=15,
        max_row=row - 1
    )

    cats = Reference(
        ws_dash,
        min_col=1,
        min_row=16,
        max_row=row - 1
    )

    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)

    bar.title = "Applications by Month"
    bar.y_axis.title = "Applications"

    ws_dash.add_chart(bar, "D18")

    #
    # Newest 10 Applications
    #

    ws_dash["J3"] = "Newest 10 Applications"

    headers = [
        "Work Type",
        "Company",
        "Position",
        "Date"
    ]

    for col, header in enumerate(headers, start=10):
        cell = ws_dash.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    row = 5

    for created_timestamp, work_type, company, position, date in newest[:10]:

        ws_dash.cell(row=row, column=10).value = work_type
        ws_dash.cell(row=row, column=11).value = company
        ws_dash.cell(row=row, column=12).value = position
        ws_dash.cell(row=row, column=13).value = date

        row += 1

    # Auto-size columns A - B
    autosize_columns(ws_dash, 1, 2)

    # Chart columns C - I
    for col in range(3, 9):  # C:I
        ws_dash.column_dimensions[get_column_letter(col)].width = 14

    # Auto-size columns J - M
    autosize_columns(ws_dash, 10, 13)

    # Date Column M width
    ws_dash.column_dimensions["M"].width = 13

    for r in range(5, row):
        ws_dash.cell(row=r, column=13).number_format = "MM/DD/YYYY"

    #
    # Freeze Header
    #

    ws_dash.freeze_panes = "A2"


def create_validation_dashboard(workbook):
    """
    Creates the Dashboard worksheet for Validation.xlsx.
    Charts are built from the History worksheet.
    """

    # Remove existing Dashboard
    if "Dashboard" in workbook.sheetnames:
        del workbook["Dashboard"]

    history = workbook["History"]
    dashboard = workbook.create_sheet("Dashboard")

    #
    # Build one record per calendar day.
    # If multiple runs occur on the same day,
    # keep only the LAST run.
    #

    daily_history = {}

    for row in history.iter_rows(min_row=2, values_only=True):

        run_date = row[0]

        if not run_date:
            continue

        #
        # History stores Run Date as text.
        #
        dt = datetime.strptime(
            run_date,
            "%m/%d/%Y %I:%M:%S %p"
        )

        date_key = dt.strftime("%Y-%m-%d")

        daily_history[date_key] = (
            dt,
            row[1],                    # Active PDFs
            row[4],                    # Rejected PDFs
            row[1] + row[4]            # Total Submitted
        )

    #
    # Sort by date
    #

    history_data = sorted(daily_history.values())

    #
    # helper data
    #

    dashboard["AA1"] = "Date"
    dashboard["AB1"] = "Active"
    dashboard["AC1"] = "Total"

    row = 2

    for dt, active, rejected, total in history_data:

        dashboard.cell(row=row, column=27).value = dt
        dashboard.cell(row=row, column=28).value = active
        dashboard.cell(row=row, column=29).value = total

        dashboard.cell(row=row, column=27).number_format = "mm/dd/yyyy"

        row += 1

    #
    # Dashboard title
    #

    dashboard.merge_cells("A1:H1")

    dashboard["A1"] = "Validation Dashboard"
    dashboard["A1"].font = Font(size=16, bold=True)
    dashboard["A1"].alignment = Alignment(horizontal="center")

    #
    # Summary Metrics
    #

    latest_date, latest_active, latest_rejected, latest_total = history_data[-1]

    dashboard["A3"] = "Current Active Applications"
    dashboard["B3"] = latest_active

    dashboard["A4"] = "Current Total Submitted"
    dashboard["B4"] = latest_total

    dashboard["A5"] = "History Days"
    dashboard["B5"] = len(history_data)

    dashboard["B3"].font = Font(size=14, bold=True)
    dashboard["B4"].font = Font(size=14, bold=True)
    dashboard["B5"].font = Font(size=14, bold=True)

    # Align summary values
    for cell in ("B3", "B4", "B5"):
        dashboard[cell].alignment = Alignment(horizontal="right")

    #
    # Section Titles
    #

    dashboard["A8"] = "Active Applications Over Time"
    dashboard["A25"] = "Total Applications Submitted"

    dashboard["A8"].font = Font(size=12, bold=True)
    dashboard["A25"].font = Font(size=12, bold=True)

    #
    # Active Applications Line Chart
    #

    active_chart = LineChart()

    data = Reference(
        dashboard,
        min_col=28,      # AB
        min_row=1,
        max_row=row - 1
    )

    dates = Reference(
        dashboard,
        min_col=27,      # AA
        min_row=2,
        max_row=row - 1
    )

    active_chart.add_data(data, titles_from_data=True)
    active_chart.set_categories(dates)

    active_chart.title = "Active Applications Over Time"
    active_chart.y_axis.title = "Applications"
    active_chart.x_axis.title = "Date"

    active_chart.height = 8
    active_chart.width = 16

    active_chart.legend = None

    dashboard.add_chart(active_chart, "D3")

    #
    # Total Submitted Line Chart
    #

    total_chart = LineChart()

    data = Reference(
        dashboard,
        min_col=29,      # AC
        min_row=1,
        max_row=row - 1
    )

    dates = Reference(
        dashboard,
        min_col=27,
        min_row=2,
        max_row=row - 1
    )

    total_chart.add_data(data, titles_from_data=True)
    total_chart.set_categories(dates)

    total_chart.title = "Total Applications Submitted"
    total_chart.y_axis.title = "Applications"
    total_chart.x_axis.title = "Date"

    total_chart.height = 8
    total_chart.width = 16

    total_chart.legend = None

    dashboard.add_chart(total_chart, "D20")

    # Hide helper columns N - Z so the AA:AC helper data used by the
    # charts sits visually right after the chart area, without being
    # hidden itself.
    for col in range(14, 27):  # N through Z
        dashboard.column_dimensions[get_column_letter(col)].hidden = True

    # Auto-size columns A - B, AB - AC
    autosize_columns(dashboard, 1, 2)
    autosize_columns(dashboard, 28, 29)

    # Make Date width smaller
    dashboard.column_dimensions["AA"].width = 12

    # Freeze header row
    dashboard.freeze_panes = "A2"

    # Optional: remove worksheet gridlines
    # dashboard.sheet_view.showGridLines = False
