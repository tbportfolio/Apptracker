"""Cross-check generated output files and maintain Validation.xlsx."""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from xlsx_utils import autosize_columns
from dashboards import create_validation_dashboard


def validate_data(
    active_pdf_count,
    rejected_pdf_count,
    excel_file,
    active_csv,
    rejected_csv
):

    wb = load_workbook(excel_file, data_only=True)

    active_sheet = wb["Active"]
    rejected_sheet = wb["Rejected"]

    active_excel = active_sheet.max_row - 1
    rejected_excel = rejected_sheet.max_row - 1

    active_csv_count = len(pd.read_csv(active_csv))
    rejected_csv_count = len(pd.read_csv(rejected_csv))

    validation = {

        "active_pdf": active_pdf_count,
        "active_excel": active_excel,
        "active_csv": active_csv_count,

        "rejected_pdf": rejected_pdf_count,
        "rejected_excel": rejected_excel,
        "rejected_csv": rejected_csv_count

    }

    validation["passed"] = (

        validation["active_pdf"] ==
        validation["active_excel"] ==
        validation["active_csv"]

        and

        validation["rejected_pdf"] ==
        validation["rejected_excel"] ==
        validation["rejected_csv"]

    )

    return validation


def save_validation(validation, validation_file, version, elapsed_time):
    """
    elapsed_time is passed in explicitly (seconds, from Audit.elapsed_seconds())
    instead of being recomputed here off a module-level start_time global.
    """

    if validation_file.exists():
        wb = load_workbook(validation_file)

    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    # -----------------------------
    # Summary Sheet
    # -----------------------------
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb.create_sheet("Summary", 0)

    ws.append(["Metric", "Value"])

    ws.append(["Run Date", datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")])

    ws.append([])

    ws.append(["Active PDFs", validation["active_pdf"]])
    ws.append(["Active Excel", validation["active_excel"]])
    ws.append(["Active CSV", validation["active_csv"]])

    ws.append([])

    ws.append(["Rejected PDFs", validation["rejected_pdf"]])
    ws.append(["Rejected Excel", validation["rejected_excel"]])
    ws.append(["Rejected CSV", validation["rejected_csv"]])

    ws.append([])

    ws.append(["Validation", "PASSED" if validation["passed"] else "FAILED"])
    ws.append(["Version", version])

    ws.append([])

    ws.append(["Elapsed Time", f"{elapsed_time:.4f} seconds"])

    ws["B2"].alignment = Alignment(horizontal="right")
    ws["B12"].alignment = Alignment(horizontal="right")
    ws["B13"].alignment = Alignment(horizontal="right")
    ws["B15"].alignment = Alignment(horizontal="right")

    # -----------------------------
    # History Sheet
    # -----------------------------
    history = wb["History"] if "History" in wb.sheetnames else wb.create_sheet("History")

    if history.max_row == 1 and history["A1"].value is None:

        history["A1"] = "Run Date"
        history["B1"] = "Active PDFs"
        history["C1"] = "Active Excel"
        history["D1"] = "Active CSV"
        history["E1"] = "Rejected PDFs"
        history["F1"] = "Rejected Excel"
        history["G1"] = "Rejected CSV"
        history["H1"] = "Validation"
        history["I1"] = "Version"
        history["J1"] = "Elapsed Time"

    history.append([

        datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),

        validation["active_pdf"],
        validation["active_excel"],
        validation["active_csv"],

        validation["rejected_pdf"],
        validation["rejected_excel"],
        validation["rejected_csv"],

        "PASSED" if validation["passed"] else "FAILED",

        version,

        round(elapsed_time, 4)

    ])

    # -----------------------------
    # Format Work Sheets
    # -----------------------------
    # Apply custom number format to Column J (Elapsed Time)
    last_row = history.max_row
    history[f"J{last_row}"].number_format = '0.0000'

    # Auto-size columns
    autosize_columns(ws)
    autosize_columns(history)

    ws.freeze_panes = "A2"
    history.freeze_panes = "A2"

    bold = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold

    for cell in history[1]:
        cell.font = bold

    # Create Validation Dashboard
    create_validation_dashboard(wb)

    wb.save(validation_file)
