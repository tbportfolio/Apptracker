#!/usr/bin/env python3
# ============================================================
# Apptracker Demo Data Generator
#
# File: make_demo_data.py
# Version: 1.2
#
# Purpose:
#     Creates Apptracker_Demo.xlsx by replacing Company and
#     Position values with realistic demo data while preserving
#     workbook formatting, hyperlinks, filters, tables, charts,
#     and all other information.
#
# Input:
#     Apptracker.xlsx
#
# Output:
#     Apptracker_Demo.xlsx
#
# Requirements:
#     pip install openpyxl
#
# Compatible with:
#     Apptracker v5.x
# ============================================================

from pathlib import Path
from openpyxl import load_workbook
import random
import sys

# ------------------------------------------------------------
# Files
# ------------------------------------------------------------

INPUT_FILE = "Apptracker.xlsx"
OUTPUT_FILE = "Apptracker_Demo.xlsx"

# ------------------------------------------------------------
# Random generator
# ------------------------------------------------------------

rng = random.Random()

# ------------------------------------------------------------
# Fake Company Names
#
# Feel free to expand this list.
# The generator randomly selects one company for every row.
# ------------------------------------------------------------

COMPANIES = [

    # ------------------------
    # Big Tech
    # ------------------------

    "Microsoft",
    "Google",
    "Apple",
    "Amazon",
    "Meta",
    "OpenAI",
    "NVIDIA",
    "Intel",
    "AMD",
    "Cisco",
    "Oracle",
    "IBM",
    "Adobe",
    "Salesforce",
    "ServiceNow",
    "Snowflake",
    "Datadog",
    "MongoDB",
    "Cloudflare",
    "Dropbox",
    "Box",
    "Asana",
    "Atlassian",
    "HubSpot",
    "Zoom",
    "Twilio",
    "Stripe",
    "PayPal",
    "Square",
    "Block",
    "Airbnb",
    "Uber",
    "Lyft",
    "DoorDash",
    "Instacart",
    "Spotify",
    "Canva",
    "Reddit",
    "Pinterest",
    "GitHub",

    # ------------------------
    # Fortune 500
    # ------------------------

    "Walmart",
    "Costco",
    "Target",
    "Home Depot",
    "Lowe's",
    "Johnson & Johnson",
    "Procter & Gamble",
    "PepsiCo",
    "Coca-Cola",
    "Nike",
    "General Electric",
    "3M",
    "Honeywell",
    "Boeing",
    "Lockheed Martin",
    "RTX",
    "Northrop Grumman",
    "Ford",
    "General Motors",
    "Tesla",

    # ------------------------
    # Finance
    # ------------------------

    "JPMorgan Chase",
    "Goldman Sachs",
    "Morgan Stanley",
    "Citigroup",
    "Bank of America",
    "Wells Fargo",
    "Capital One",
    "American Express",
    "Mastercard",
    "Visa",
    "BlackRock",
    "Fidelity",
    "Charles Schwab",
    "State Street",
    "Bloomberg",

    # ------------------------
    # Healthcare
    # ------------------------

    "Pfizer",
    "Moderna",
    "AbbVie",
    "Merck",
    "Bristol Myers Squibb",
    "UnitedHealth Group",
    "CVS Health",
    "Cigna",
    "Elevance Health",
    "Mayo Clinic",
    "Cleveland Clinic",
    "Mount Sinai Health System",
    "NewYork-Presbyterian",
    "Northwell Health",
    "Memorial Sloan Kettering",

    # ------------------------
    # Government
    # ------------------------

    "NASA",
    "FEMA",
    "Department of Energy",
    "Department of Transportation",
    "Department of Veterans Affairs",
    "National Institutes of Health",
    "New York State ITS",
    "New York State Department of Health",
    "City of New York",
    "Port Authority of New York & New Jersey",

    # ------------------------
    # Universities
    # ------------------------

    "Cornell University",
    "Columbia University",
    "New York University",
    "Princeton University",
    "Yale University",
    "Harvard University",
    "Massachusetts Institute of Technology",
    "Stanford University",
    "University of Michigan",
    "Penn State University",

    # ------------------------
    # Consulting
    # ------------------------

    "Accenture",
    "Deloitte",
    "PwC",
    "EY",
    "KPMG",
    "McKinsey & Company",
    "Boston Consulting Group",
    "Bain & Company",
    "Capgemini",
    "Slalom",

    # ------------------------
    # Telecom
    # ------------------------

    "AT&T",
    "Verizon",
    "T-Mobile",
    "Comcast",
    "Charter Communications",

    # ------------------------
    # Energy
    # ------------------------

    "ExxonMobil",
    "Chevron",
    "ConocoPhillips",
    "NextEra Energy",
    "Duke Energy",

    # ------------------------
    # Retail / Consumer
    # ------------------------

    "Best Buy",
    "Macy's",
    "Kohl's",
    "Starbucks",
    "McDonald's",
    "Chipotle",
    "Domino's Pizza",
    "Hilton",
    "Marriott",
    "Delta Air Lines"

]


# ------------------------------------------------------------
# Fake Job Titles
#
# Approximately 120 modern business, program management,
# operations, analytics, AI, and technology titles.
# ------------------------------------------------------------

JOB_TITLES = [

    # --------------------------------------------------------
    # Program Management
    # --------------------------------------------------------

    "Program Manager",
    "Senior Program Manager",
    "Principal Program Manager",
    "Lead Program Manager",
    "Enterprise Program Manager",
    "Technical Program Manager",
    "Senior Technical Program Manager",
    "Principal Technical Program Manager",
    "Strategic Program Manager",
    "Transformation Program Manager",
    "Portfolio Program Manager",
    "PMO Program Manager",
    "Program Delivery Manager",
    "Business Program Manager",
    "Global Program Manager",

    # --------------------------------------------------------
    # Project Management
    # --------------------------------------------------------

    "Project Manager",
    "Senior Project Manager",
    "Principal Project Manager",
    "Implementation Project Manager",
    "Project Management Office Manager",
    "Infrastructure Project Manager",
    "Business Project Manager",
    "Project Portfolio Manager",
    "Capital Projects Manager",
    "PMO Manager",

    # --------------------------------------------------------
    # Operations
    # --------------------------------------------------------

    "Operations Manager",
    "Senior Operations Manager",
    "Business Operations Manager",
    "Director of Operations",
    "Head of Operations",
    "Operations Excellence Manager",
    "Operational Excellence Manager",
    "Operations Strategy Manager",
    "Operations Director",
    "Operations Analyst",
    "Operations Consultant",
    "Business Operations Analyst",
    "Process Operations Manager",
    "Enterprise Operations Manager",
    "Regional Operations Manager",

    # --------------------------------------------------------
    # Product
    # --------------------------------------------------------

    "Product Manager",
    "Senior Product Manager",
    "Principal Product Manager",
    "Lead Product Manager",
    "Technical Product Manager",
    "Product Operations Manager",
    "Digital Product Manager",
    "Platform Product Manager",
    "AI Product Manager",
    "Director of Product Management",

    # --------------------------------------------------------
    # Business Analysis
    # --------------------------------------------------------

    "Business Analyst",
    "Senior Business Analyst",
    "Lead Business Analyst",
    "Business Systems Analyst",
    "Senior Business Systems Analyst",
    "Business Systems Manager",
    "Business Process Analyst",
    "Business Process Manager",
    "Business Intelligence Analyst",
    "Business Intelligence Manager",

    # --------------------------------------------------------
    # Strategy
    # --------------------------------------------------------

    "Strategy Manager",
    "Senior Strategy Manager",
    "Corporate Strategy Manager",
    "Business Strategy Manager",
    "Strategic Planning Manager",
    "Director of Strategy",
    "Strategy Consultant",
    "Transformation Manager",

    # --------------------------------------------------------
    # Analytics / Data
    # --------------------------------------------------------

    "Analytics Manager",
    "Senior Analytics Manager",
    "Data Analytics Manager",
    "Data Program Manager",
    "Data Operations Manager",
    "Reporting Manager",
    "Reporting Analyst",
    "Business Intelligence Developer",
    "SQL Developer",
    "Data Warehouse Manager",
    "Data Governance Manager",
    "Data Quality Manager",

    # --------------------------------------------------------
    # Digital Transformation
    # --------------------------------------------------------

    "Digital Transformation Manager",
    "Enterprise Transformation Manager",
    "Business Transformation Manager",
    "Process Improvement Manager",
    "Continuous Improvement Manager",
    "Lean Process Manager",
    "Operational Improvement Manager",
    "Transformation Director",

    # --------------------------------------------------------
    # AI
    # --------------------------------------------------------

    "AI Transformation Lead",
    "AI Program Manager",
    "AI Operations Manager",
    "AI Solutions Manager",
    "Generative AI Program Manager",
    "AI Business Consultant",
    "AI Implementation Manager",
    "AI Strategy Manager",
    "AI Enablement Lead",
    "Automation Program Manager",

    # --------------------------------------------------------
    # Technology
    # --------------------------------------------------------

    "Technology Manager",
    "Technology Program Manager",
    "Solutions Architect",
    "Solutions Consultant",
    "Technical Solutions Manager",
    "Enterprise Applications Manager",
    "Application Support Manager",
    "Systems Integration Manager",
    "Technology Consultant",
    "Information Systems Manager",

    # --------------------------------------------------------
    # Consulting
    # --------------------------------------------------------

    "Management Consultant",
    "Senior Consultant",
    "Principal Consultant",
    "Business Consultant",
    "Transformation Consultant",
    "Operations Consultant",
    "Strategy Consultant",
    "Program Management Consultant",
    "Technology Consultant",

    # --------------------------------------------------------
    # Leadership
    # --------------------------------------------------------

    "Chief of Staff",
    "Director",
    "Senior Director",
    "Executive Director",
    "Vice President",
    "Managing Director",
    "Portfolio Director",
    "Program Director",
    "Operations Director",
    "Business Unit Director",

    # --------------------------------------------------------
    # Miscellaneous
    # --------------------------------------------------------

    "Customer Success Operations Manager",
    "Implementation Manager",
    "Implementation Consultant",
    "Client Success Manager",
    "Vendor Management Manager",
    "Risk Program Manager",
    "Compliance Program Manager",
    "Governance Manager",
    "Change Management Lead",
    "Organizational Change Manager"

]


# ------------------------------------------------------------
# Utility Constants
# ------------------------------------------------------------

TARGET_SHEETS = [
    "Active",
    "Rejected"
]

TARGET_COLUMNS = {
    "company": None,
    "position": None
}

HEADER_ROW = 1

BANNER = "=" * 60


# ------------------------------------------------------------
# Helper Functions
# ------------------------------------------------------------

def print_banner():
    """Display program banner."""

    print()
    print(BANNER)
    print("Apptracker Demo Data Generator v1.2")
    print(BANNER)
    print()


def random_company():
    """Return a random fake company."""

    return rng.choice(COMPANIES)


def random_job_title():
    """Return a random job title."""

    return rng.choice(JOB_TITLES)


def file_exists(filename):
    """Verify the input workbook exists."""

    if not Path(filename).exists():
        print(f"ERROR: '{filename}' not found.")
        print()
        sys.exit(1)


def find_column_indexes(ws):
    """
    Locate the Company and Position columns by
    reading the worksheet headers.

    Returns:
        dict containing:
            company
            position
    """

    columns = {
        "company": None,
        "position": None
    }

    for cell in ws[HEADER_ROW]:

        if cell.value is None:
            continue

        header = str(cell.value).strip().lower()

        if header == "company":
            columns["company"] = cell.column

        elif header == "position":
            columns["position"] = cell.column

    missing = []

    if columns["company"] is None:
        missing.append("Company")

    if columns["position"] is None:
        missing.append("Position")

    if missing:

        raise ValueError(
            f"{ws.title}: Missing required column(s): "
            + ", ".join(missing)
        )

    return columns


def get_target_sheets(workbook):
    """
    Return all sheets that should be randomized.
    """

    sheets = []

    for sheet_name in TARGET_SHEETS:

        if sheet_name in workbook.sheetnames:
            sheets.append(workbook[sheet_name])

    return sheets


def replace_row_values(ws, row, columns):
    """
    Replace Company and Position values
    for a single row.
    """

    company_cell = ws.cell(row=row,
                           column=columns["company"])

    position_cell = ws.cell(row=row,
                            column=columns["position"])

    if company_cell.value not in (None, ""):
        company_cell.value = random_company()

    if position_cell.value not in (None, ""):
        position_cell.value = random_job_title()


def count_data_rows(ws):
    """
    Count rows containing application data.
    """

    count = 0

    for row in range(2, ws.max_row + 1):

        values = [
            ws.cell(row=row, column=col).value
            for col in range(1, ws.max_column + 1)
        ]

        if any(value not in (None, "") for value in values):
            count += 1

    return count


def save_workbook(workbook):
    """
    Save the randomized workbook.
    """

    workbook.save(OUTPUT_FILE)


# ------------------------------------------------------------
# Workbook Processing Functions
# ------------------------------------------------------------

def process_sheet(ws):
    """
    Randomize Company and Position columns on a worksheet.

    Returns:
        Number of rows processed.
    """

    print(f"Processing worksheet: {ws.title}")

    columns = find_column_indexes(ws)

    processed = 0

    for row in range(2, ws.max_row + 1):

        values = [
            ws.cell(row=row, column=col).value
            for col in range(1, ws.max_column + 1)
        ]

        if not any(value not in (None, "") for value in values):
            continue

        replace_row_values(ws, row, columns)

        processed += 1

    print(f"   Rows processed : {processed}")

    return processed


def process_workbook(workbook):
    """
    Process every target worksheet.

    Returns:
        Total rows processed.
    """

    total_rows = 0

    worksheets = get_target_sheets(workbook)

    if not worksheets:
        raise ValueError(
            "No Active or Rejected worksheet found."
        )

    print()

    for ws in worksheets:

        total_rows += process_sheet(ws)

        print()

    return total_rows


def update_dashboard(workbook):
    """
    Update the Dashboard's Newest 10 Applications
    using the newest rows from the Active sheet.
    """

    if "Dashboard" not in workbook.sheetnames:
        return

    if "Active" not in workbook.sheetnames:
        return

    dashboard = workbook["Dashboard"]
    active = workbook["Active"]

    headers = {}

    for cell in active[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column

    company_col = headers["company"]
    position_col = headers["position"]
    created_col = headers["created timestamp"]

    rows = []

    for row in range(2, active.max_row + 1):

        created = active.cell(row=row, column=created_col).value

        if created is None:
            continue

        rows.append((created, row))

    rows.sort(reverse=True)

    newest_rows = rows[:10]

    dashboard_row = 5

    for _, active_row in newest_rows:

        dashboard.cell(
            row=dashboard_row,
            column=11
        ).value = active.cell(
            row=active_row,
            column=company_col
        ).value

        dashboard.cell(
            row=dashboard_row,
            column=12
        ).value = active.cell(
            row=active_row,
            column=position_col
        ).value

        dashboard_row += 1


def create_demo_workbook():
    """
    Main workbook processing routine.

    Opens Apptracker.xlsx,
    randomizes sensitive fields,
    and saves Apptracker_Demo.xlsx.

    Returns:
        Number of rows processed.
    """

    file_exists(INPUT_FILE)

    print(f"Loading workbook: {INPUT_FILE}")

    workbook = load_workbook(INPUT_FILE)

    print("Workbook loaded successfully.")
    print()

    total_rows = process_workbook(workbook)

    update_dashboard(workbook)

    print("Saving workbook...")

    save_workbook(workbook)

    print("Workbook saved.")
    print()

    return total_rows


def print_summary(total_rows):
    """
    Display completion summary.
    """

    print(BANNER)
    print("Demo workbook created successfully.")
    print(BANNER)
    print()

    print(f"Input Workbook : {INPUT_FILE}")
    print(f"Output Workbook: {OUTPUT_FILE}")
    print(f"Rows Processed : {total_rows}")

    print()
    print("Company and Position fields were randomized.")
    print("All other workbook data was preserved.")
    print()


# ------------------------------------------------------------
# Main Program
# ------------------------------------------------------------

def main():
    """
    Main program entry point.
    """

    print_banner()

    try:

        total_rows = create_demo_workbook()

        print_summary(total_rows)

    except KeyboardInterrupt:

        print()
        print("Program cancelled by user.")
        sys.exit(1)

    except Exception as ex:

        print()
        print(BANNER)
        print("ERROR")
        print(BANNER)
        print(ex)
        print()

        sys.exit(1)


# ------------------------------------------------------------
# Program Entry Point
# ------------------------------------------------------------

if __name__ == "__main__":
    main()